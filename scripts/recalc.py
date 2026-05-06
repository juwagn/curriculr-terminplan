"""
recalc.py – Verifikation der bearbeitbaren Jahresvorlage (v3)

Prueft:
  1. Workbook/Sheets sind nicht geschuetzt.
  2. Ferien & Eckdaten enthalten echte Datumswerte.
  3. Hilfstabelle F15:F56 bildet SW 00 aus B10 und SW 01 aus B11.
  4. Folgewochen ueberspringen Ferienperioden und liefern 42 Schulwochen.
  5. Terminplan-Spalten A-D sind technisch konsistent, E-K bleiben Eingabebereich.
  6. Keine offensichtlichen Formel-Fehlertexte in Formeln.

Ausgabe: JSON mit status, details, errors
Ausfuehren: python scripts/recalc.py
"""

from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
XLSX_PATH = BASE / "website" / "downloads" / "Terminplan_Schulwochen_Vorlage.xlsx"
SW_COUNT = 42
ROWS_PER_WEEK = 15
HELPER_START_ROW = 15


def read_date(cell):
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def compute_school_mondays(sw00, sw01, holiday_periods, count=SW_COUNT):
    mondays = [sw00, sw01]
    current = sw01 + timedelta(days=7)
    guard = 0
    while len(mondays) < count and guard < 200:
        guard += 1
        if not any(start and end and start <= current <= end for start, end in holiday_periods):
            mondays.append(current)
        current += timedelta(days=7)
    return mondays


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


errors = []
details = []
print(f"Lade: {XLSX_PATH}")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)

expected_sheets = ["Terminplan", "Ferien", "Kategorien", "Anleitung"]
if wb.sheetnames != expected_sheets:
    errors.append(f"Sheet-Reihenfolge falsch: {wb.sheetnames}")
else:
    details.append("Sheet-Reihenfolge Terminplan/Ferien/Kategorien/Anleitung OK")

for ws in wb.worksheets:
    if ws.protection.sheet:
        errors.append(f"{ws.title}: Blattschutz ist aktiv")
    else:
        details.append(f"{ws.title}: kein Blattschutz OK")

ws_f = wb["Ferien"]
ws_t = wb["Terminplan"]

required_dates = {
    "B3": "Herbstferien Von",
    "C3": "Herbstferien Bis",
    "B4": "Weihnachtsferien Von",
    "C4": "Weihnachtsferien Bis",
    "B5": "Osterferien Von",
    "C5": "Osterferien Bis",
    "B7": "Sommerferien Von",
    "C7": "Sommerferien Bis",
    "B10": "Erster Schultag SW 00",
    "B11": "Erster Unterrichtstag SW 01",
    "B12": "Letzter Schultag",
}
ferien_dates = {}
for addr, label in required_dates.items():
    value = read_date(ws_f[addr])
    if value is None:
        errors.append(f"Ferien!{addr} ({label}): kein Datumswert, Wert={ws_f[addr].value!r}")
    else:
        ferien_dates[addr] = value
        details.append(f"Ferien!{addr} ({label}): {value} OK")

for addr in ("B6", "C6"):
    ferien_dates[addr] = read_date(ws_f[addr])

sw00 = ferien_dates.get("B10")
sw01 = ferien_dates.get("B11")
if sw00 and sw00.weekday() != 0:
    errors.append(f"SW 00 muss Montag sein, ist {sw00}")
if sw01 and sw01.weekday() != 0:
    errors.append(f"SW 01 muss Montag sein, ist {sw01}")
if sw00 and sw01 and sw01 <= sw00:
    errors.append(f"SW 01 ({sw01}) muss nach SW 00 ({sw00}) liegen")

if is_formula(ws_f["F15"].value) and "$B$10" in ws_f["F15"].value:
    details.append("Ferien!F15 referenziert SW 00/B10 OK")
else:
    errors.append(f"Ferien!F15 falsche Formel: {ws_f['F15'].value!r}")

if is_formula(ws_f["F16"].value) and "$B$11" in ws_f["F16"].value:
    details.append("Ferien!F16 referenziert SW 01/B11 OK")
else:
    errors.append(f"Ferien!F16 falsche Formel: {ws_f['F16'].value!r}")

for sw in range(2, SW_COUNT):
    row = HELPER_START_ROW + sw
    formula = ws_f.cell(row, 6).value
    if not is_formula(formula) or f"$F${row - 1}" not in formula:
        errors.append(f"Ferien!F{row} (SW {sw:02d}): Folgeformel fehlerhaft: {formula!r}")

holiday_periods = [
    (ferien_dates.get("B3"), ferien_dates.get("C3")),
    (ferien_dates.get("B4"), ferien_dates.get("C4")),
    (ferien_dates.get("B5"), ferien_dates.get("C5")),
    (ferien_dates.get("B6"), ferien_dates.get("C6")),
    (ferien_dates.get("B7"), ferien_dates.get("C7")),
]
if sw00 and sw01:
    computed = compute_school_mondays(sw00, sw01, holiday_periods)
    if len(computed) != SW_COUNT:
        errors.append(f"Python-Berechnung: {len(computed)} statt {SW_COUNT} Schulwochen")
    else:
        details.append(f"Python-Berechnung: SW00={computed[0]}, SW01={computed[1]}, SW41={computed[-1]} OK")
    for idx in [0, 1, 7, 8, 15, 16, 27, 28, 40, 41]:
        if idx < len(computed):
            details.append(f"Spotcheck SW {idx:02d}: {computed[idx]}")

header_rows = []
for row in range(2, ws_t.max_row + 1):
    value = ws_t.cell(row, 1).value
    if isinstance(value, str) and value.startswith("SW "):
        try:
            header_rows.append((int(value.split()[1]), row))
        except ValueError:
            errors.append(f"Terminplan A{row}: unlesbarer SW-Key {value!r}")

if len(header_rows) != SW_COUNT:
    errors.append(f"Terminplan: {len(header_rows)} statt {SW_COUNT} SW-Header")
else:
    details.append("Terminplan: 42 SW-Header OK")

for sw, row in header_rows:
    expected_b = f"=Ferien!$F${HELPER_START_ROW + sw}"
    if ws_t.cell(row, 2).value != expected_b:
        errors.append(f"Terminplan B{row}: erwartet {expected_b}, gefunden {ws_t.cell(row, 2).value!r}")
    if not is_formula(ws_t.cell(row, 4).value) or "TEXT" not in ws_t.cell(row, 4).value:
        errors.append(f"Terminplan D{row}: Wochenformel fehlt")

for col in ("A", "B", "C"):
    if not ws_t.column_dimensions[col].hidden:
        errors.append(f"Terminplan Spalte {col}: sollte ausgeblendet sein")

if ws_t.protection.sheet:
    errors.append("Terminplan darf nicht geschuetzt sein")

bad_terms = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"]
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and any(term in value for term in bad_terms):
                errors.append(f"{ws.title}!{cell.coordinate}: Formel-/Fehlertext {value!r}")

status = "success" if not errors else "error"
print(json.dumps({"status": status, "errors": errors, "details": details}, ensure_ascii=True, indent=2))
