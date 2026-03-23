# -*- coding: utf-8 -*-
"""
patch_rows.py
Baut die Terminplan-Grundstruktur neu auf:
- SW 00-41 (42 Schulwochen) x 21 Zeilen (1 Header + 20 Datenzeilen)
- Gesamt: 1 Kopfzeile + 882 Folgezeilen = 883 Zeilen

Muss vor patch_xlsx.py und build_excel_template.py ausgefuehrt werden.
Bei Strukturaenderungen (andere Zeilenanzahl, mehr SW) neu ausfuehren.
"""

from pathlib import Path
import openpyxl

BASE = Path(__file__).resolve().parent.parent
XLSX = BASE / 'website' / 'downloads' / 'Terminplan_Schulwochen_Vorlage.xlsx'

ROWS_PER_SW = 20   # Datenzeilen pro Schulwoche (ohne SW-Header-Zeile)
SW_COUNT    = 42   # SW 00-41


def rebuild_terminplan(ws):
    """Loescht Terminplan-Inhalt ab Zeile 2 und schreibt neue SW-Struktur."""
    # Alle Zeilen ab Zeile 2 loeschen (Zeile 1 = Spalten-Header bleibt)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # SW-Header-Zeilen schreiben; Datenzeilen bleiben leer
    current_row = 2
    for sw_num in range(SW_COUNT):
        ws.cell(row=current_row, column=1).value = f'SW {sw_num:02d}'
        current_row += ROWS_PER_SW + 1   # 1 Header + 15 Daten -> naechster Header

    # Letzte Datenzeile mit leerem String schreiben damit max_row = 883 ist.
    # openpyxl zaehlt nur Zeilen mit Zellinhalten; patch_xlsx.py und
    # build_excel_template.py nutzen max_row fuer Dropdown-Ranges und Styling.
    last_data_row = 1 + SW_COUNT * (ROWS_PER_SW + 1)   # = 883
    ws.cell(row=last_data_row, column=4).value = ''


print(f'Lade: {XLSX}')
wb = openpyxl.load_workbook(XLSX)

ws_t = wb['Terminplan']
rebuild_terminplan(ws_t)

last_sw_header_row = 2 + (SW_COUNT - 1) * (ROWS_PER_SW + 1)   # = 863
print(f'  SW 00: Zeile 2')
print(f'  SW 41: Zeile {last_sw_header_row}')
print(f'  max_row: {ws_t.max_row}')

wb.save(XLSX)
print(f'Gespeichert: {XLSX}')
print('Weiter mit: python scripts/patch_xlsx.py')
