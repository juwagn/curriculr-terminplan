# -*- coding: utf-8 -*-
"""
build_excel_template.py

Generiert die bearbeitbare GSH-Schulwochen-Vorlage neu.

Prinzip:
- Keine Blatt- oder Arbeitsmappensperre. Die Schulleitung kann die Datei jedes Jahr frei anpassen.
- Ferien & Eckdaten sind die einzige zentrale Eingabestelle fuer Schuljahresdaten.
- SW 00 wird aus "Erster Schultag (SW 00)" berechnet.
- SW 01 wird aus "Erster Unterrichtstag (SW 01)" berechnet.
- Folgewochen springen automatisch ueber eingetragene Ferienperioden.
- Der Konverter kann die ausgeblendeten technischen Spalten A-C lesen.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

BASE = Path(__file__).resolve().parent.parent
OUT = BASE / "website" / "downloads" / "Terminplan_Schulwochen_Vorlage.xlsx"

SW_COUNT = 42
ROWS_PER_WEEK = 15
HELPER_START_ROW = 15

# Voreinstellung 2026/27; im Ferien-Sheet frei anpassbar.
DEFAULTS = {
    "herbst_start": date(2026, 10, 19),
    "herbst_end": date(2026, 10, 30),
    "weihnachten_start": date(2026, 12, 22),
    "weihnachten_end": date(2027, 1, 3),
    "ostern_start": date(2027, 3, 22),
    "ostern_end": date(2027, 4, 2),
    "pfingsten_start": None,
    "pfingsten_end": None,
    "sommer_start": date(2027, 7, 19),
    "sommer_end": date(2027, 8, 31),
    "first_school_day": date(2026, 8, 24),
    "first_lesson_day": date(2026, 8, 31),
    "last_school_day": date(2027, 7, 16),
}

C_HEADER_BG = "1E3A5F"
C_HEADER_2 = "2F5496"
C_INPUT = "FFF7D6"
C_INPUT_DARK = "B7791F"
C_INFO = "EEF2FF"
C_INFO_TEXT = "3730A3"
C_WHITE = "FFFFFF"
C_ZEBRA = "F8FAFC"
C_BORDER = "D0D5DD"
C_TECH = "E5E7EB"
C_WEEK = "4F46E5"
C_WEEK_SW00 = "047857"
C_WARN = "FEE2E2"
C_WARN_TEXT = "991B1B"

CAT_COLORS = {
    "Jahrgang 5/6": ("FADBD8", "7B241C"),
    "Jahrgang 7/8": ("D5F5E3", "1E8449"),
    "Jahrgang 9/10": ("FDE8D3", "784212"),
    "Oberstufe": ("D4E6F1", "1A5276"),
    "Inklusion": ("D1F2EB", "148F77"),
    "Feiertage/Ferien": ("FDEBD0", "784212"),
    "Konferenzen/DB": ("E8DAEF", "6C3483"),
}

CAT_EXAMPLES = {
    "Jahrgang 5/6": "Einschulung, Sprachstandstest, Neue 5er",
    "Jahrgang 7/8": "Potenzialanalyse, WP-Wahl, KAoA",
    "Jahrgang 9/10": "Betriebspraktikum, ZP 10, Abschlussfahrt",
    "Oberstufe": "Abitur, EF/Q1/Q2 Termine",
    "Inklusion": "IFOe, AL SuS, Inklusionsteam",
    "Feiertage/Ferien": "Ferien, Feiertage, schulfreie Tage",
    "Konferenzen/DB": "Lehrerkonferenz, FaKo, Teamgespraeche",
}


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def thin_border() -> Border:
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def set_date(cell, value):
    cell.value = value
    cell.number_format = "DD.MM.YYYY"


def week_formula(prev_row: int) -> str:
    fp = f"$F${prev_row}"

    def skip(start: str, end: str) -> str:
        return (
            f"+IF(ISBLANK({start}),0,"
            f"IF(({fp}+7>={start})*({fp}+7<={end}),"
            f"{end}+8-WEEKDAY({end},2)-({fp}+7),0))"
        )

    return (
        f"={fp}+7"
        + skip("$B$3", "$C$3")
        + skip("$B$4", "$C$4")
        + skip("$B$5", "$C$5")
        + skip("$B$6", "$C$6")
        + skip("$B$7", "$C$7")
    )


def make_ferien(wb: Workbook):
    ws = wb.create_sheet("Ferien")
    ws.sheet_properties.tabColor = "ED7D31"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    widths = {"A": 34, "B": 16, "C": 16, "D": 46, "E": 12, "F": 16, "G": 28}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = '=IFERROR("Schuljahr "&TEXT(YEAR($B$10),"0")&"/"&TEXT(YEAR($B$10)+1-2000,"00")&" - Ferien & Eckdaten","Ferien & Eckdaten")'
    title.font = Font(name="Calibri", size=16, bold=True, color="1E3A5F")
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Bezeichnung", "Von (Datum)", "Bis (Datum)", "Hinweis"]
    for col, label in enumerate(headers, 1):
        cell = ws.cell(2, col, label)
        cell.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
        cell.fill = fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()

    rows = [
        (3, "Herbstferien", DEFAULTS["herbst_start"], DEFAULTS["herbst_end"], "Pflichtfeld"),
        (4, "Weihnachtsferien", DEFAULTS["weihnachten_start"], DEFAULTS["weihnachten_end"], "Pflichtfeld"),
        (5, "Osterferien", DEFAULTS["ostern_start"], DEFAULTS["ostern_end"], "Pflichtfeld"),
        (6, "Pfingstferien", DEFAULTS["pfingsten_start"], DEFAULTS["pfingsten_end"], "Optional leer lassen"),
        (7, "Sommerferien", DEFAULTS["sommer_start"], DEFAULTS["sommer_end"], "Pflichtfeld"),
    ]
    for row, name, start, end, note in rows:
        bg = C_ZEBRA if row % 2 == 0 else C_WHITE
        ws.cell(row, 1, name)
        if start:
            set_date(ws.cell(row, 2), start)
        if end:
            set_date(ws.cell(row, 3), end)
        ws.cell(row, 4, note)
        for col in range(1, 5):
            c = ws.cell(row, col)
            c.fill = fill(C_INPUT if col in (2, 3) else bg)
            c.font = Font(name="Calibri", size=11, color=(C_INPUT_DARK if col in (2, 3) else "111827"), bold=(col == 1))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border()

    ws.cell(9, 1, "Schuljahres-Eckdaten")
    ws.cell(9, 1).font = Font(name="Calibri", size=11, bold=True, color=C_INFO_TEXT)

    info_rows = [
        (10, "Erster Schultag (SW 00)", DEFAULTS["first_school_day"], "SW 00: Vorbereitungstage, Konferenzen, Schulstart vor Unterricht"),
        (11, "Erster Unterrichtstag (SW 01)", DEFAULTS["first_lesson_day"], "SW 01: erster regulaerer Unterrichtstag"),
        (12, "Letzter Schultag", DEFAULTS["last_school_day"], "Kontrollwert; Terminplan markiert letzte Woche"),
    ]
    for row, label, value, note in info_rows:
        ws.cell(row, 1, label)
        set_date(ws.cell(row, 2), value)
        ws.cell(row, 4, note)
        for col in range(1, 5):
            c = ws.cell(row, col)
            c.fill = fill(C_INPUT if col == 2 else (C_ZEBRA if row % 2 == 0 else C_WHITE))
            c.font = Font(name="Calibri", size=11, color=(C_INPUT_DARK if col == 2 else "111827"), bold=(col == 1))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border()

    ws.cell(14, 5, "Berechnete Schulwochen - nicht manuell pflegen")
    ws.cell(14, 5).font = Font(name="Calibri", size=10, bold=True, italic=True, color="6B7280")
    ws.cell(14, 6, "Montag")
    ws.cell(14, 7, "Hinweis")
    for col in range(5, 8):
        c = ws.cell(14, col)
        c.fill = fill(C_TECH)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center")

    for sw in range(SW_COUNT):
        row = HELPER_START_ROW + sw
        ws.cell(row, 5, f"SW {sw:02d}")
        if sw == 0:
            ws.cell(row, 6, "=$B$10")
        elif sw == 1:
            ws.cell(row, 6, "=$B$11")
        else:
            ws.cell(row, 6, week_formula(row - 1))
        ws.cell(row, 6).number_format = "DD.MM.YYYY"
        ws.cell(row, 7, '=IF(F{0}>$B$12,"nach letztem Schultag","")'.format(row))
        for col in range(5, 8):
            c = ws.cell(row, col)
            c.fill = fill(C_TECH if sw % 2 == 0 else "F3F4F6")
            c.border = thin_border()
            c.font = Font(name="Calibri", size=9, color="374151")

    for rng in ["B3:C7", "B10:B12"]:
        dv = DataValidation(type="date", allow_blank=True, showErrorMessage=True, errorTitle="Datum erwartet", error="Bitte ein echtes Excel-Datum eintragen.")
        dv.add(rng)
        ws.add_data_validation(dv)

    return ws


def make_kategorien(wb: Workbook):
    ws = wb.create_sheet("Kategorien")
    ws.sheet_properties.tabColor = "7030A0"
    ws.sheet_view.showGridLines = False
    widths = {"A": 22, "B": 14, "C": 58}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:C1")
    ws["A1"] = "Kategorien-Referenz"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
    ws.row_dimensions[1].height = 26

    for col, label in enumerate(["Kategorie", "Farbe", "Automatisch erkannte Stichwoerter"], 1):
        c = ws.cell(2, col, label)
        c.font = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
        c.fill = fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    row = 3
    for cat, (bg, fg) in CAT_COLORS.items():
        ws.cell(row, 1, cat)
        ws.cell(row, 2, f"#{bg}")
        ws.cell(row, 3, CAT_EXAMPLES.get(cat, ""))
        for col in range(1, 4):
            c = ws.cell(row, col)
            c.fill = fill(bg)
            c.font = Font(name="Calibri", size=11, bold=(col == 1), color=fg)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border()
        row += 1

    wb.defined_names.add(DefinedName("KategorienListe", attr_text=f"Kategorien!$A$3:$A${row - 1}"))
    return ws


def make_terminplan(wb: Workbook):
    ws = wb.create_sheet("Terminplan", 0)
    ws.sheet_properties.tabColor = "2F5496"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E2"

    widths = {"A": 8, "B": 14, "C": 8, "D": 24, "E": 14, "F": 10, "G": 12, "H": 44, "I": 20, "J": 12, "K": 34}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in ("A", "B", "C"):
        ws.column_dimensions[col].hidden = True

    headers = ["SW-Key", "Montag-ISO", "SW", "Schulwoche", "Wochentag", "Uhrzeit", "Endzeit", "Titel / Veranstaltung", "Kategorie", "Ganztaegig", "Anmerkung"]
    for col, label in enumerate(headers, 1):
        c = ws.cell(1, col, label)
        c.fill = fill(C_HEADER_BG)
        c.font = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[1].height = 22

    row = 2
    for sw in range(SW_COUNT):
        helper_row = HELPER_START_ROW + sw
        visible_week = f"SW {sw:02d}"
        ws.cell(row, 1, visible_week)
        ws.cell(row, 2, f"=Ferien!$F${helper_row}")
        ws.cell(row, 2).number_format = "YYYY-MM-DD"
        ws.cell(row, 3, visible_week)
        ws.cell(row, 4, f'=TEXT(B{row},"TT.MM.")&" - "&TEXT(B{row}+4,"TT.MM.JJJJ")')
        ws.cell(row, 11, f'=IF(Ferien!$G${helper_row}="", "", Ferien!$G${helper_row})')
        header_fill = C_WEEK_SW00 if sw == 0 else C_WEEK
        for col in range(1, 12):
            c = ws.cell(row, col)
            c.fill = fill(header_fill if col >= 4 else C_TECH)
            c.font = Font(name="Calibri", size=10, bold=True, color=(C_WHITE if col >= 4 else "374151"))
            c.alignment = Alignment(vertical="center", horizontal=("left" if col in (4, 11) else "center"), wrap_text=True)
            c.border = thin_border()
        ws.row_dimensions[row].height = 21
        row += 1

        for _ in range(ROWS_PER_WEEK):
            ws.cell(row, 1, None)
            ws.cell(row, 2, f"=Ferien!$F${helper_row}")
            ws.cell(row, 2).number_format = "YYYY-MM-DD"
            ws.cell(row, 3, visible_week)
            ws.cell(row, 4, None)
            bg = C_ZEBRA if row % 2 == 0 else C_WHITE
            for col in range(1, 12):
                c = ws.cell(row, col)
                c.fill = fill(bg if col >= 4 else C_TECH)
                c.font = Font(name="Calibri", size=11, color="111827")
                c.alignment = Alignment(vertical="center", wrap_text=(col in (8, 11)))
                c.border = thin_border()
            ws.row_dimensions[row].height = 18
            row += 1

    max_row = row - 1
    validations = [
        ("E2:E" + str(max_row), '"Mo,Di,Mi,Do,Fr,Ganze Woche"', "Bitte Wochentag aus Dropdown waehlen."),
        ("I2:I" + str(max_row), "KategorienListe", "Bitte Kategorie aus Dropdown waehlen."),
        ("J2:J" + str(max_row), '"Ja,Nein"', 'Bitte "Ja" oder "Nein" auswaehlen.'),
    ]
    for sqref, formula, error in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True, errorTitle="Ungueltige Eingabe", error=error)
        dv.add(sqref)
        ws.add_data_validation(dv)

    return ws


def make_anleitung(wb: Workbook):
    ws = wb.create_sheet("Anleitung")
    ws.sheet_properties.tabColor = "70AD47"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 72

    r = 1

    def write(row: int, col: int, value: str, *, bold=False, size=11, color="374151", bg=None, wrap=True):
        c = ws.cell(row, col, value)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(vertical="top", wrap_text=wrap)
        if bg:
            c.fill = fill(bg)
        return c

    def section(text: str):
        nonlocal r
        r += 1
        write(r, 2, text, bold=True, size=12, color=C_INFO_TEXT)
        r += 1

    write(2, 2, "GSH Schuljahreskalender - Anleitung", bold=True, size=18, color="1E3A5F", wrap=False)
    r = 4
    section("Jedes Jahr neu einrichten")
    steps = [
        ("1. Ferien & Eckdaten", "Im Tab 'Ferien' Ferienperioden, Erster Schultag (SW 00), Erster Unterrichtstag (SW 01) und Letzter Schultag eintragen."),
        ("2. Termine", "Im Tab 'Terminplan' nur in den sichtbaren Spalten E-K arbeiten."),
        ("3. Konverter", "Excel-Datei in Terminplan_Konverter.html ziehen und ICS-Datei erzeugen."),
        ("4. IServ", "ICS-Datei in IServ importieren."),
    ]
    for title, detail in steps:
        write(r, 2, title, bold=True, bg=C_INFO, color=C_INFO_TEXT)
        write(r, 3, detail, bg=C_INFO)
        ws.row_dimensions[r].height = None
        r += 1

    section("Wichtig")
    notes = [
        "Keine Sperre: Die Datei bleibt bewusst bearbeitbar. Technische Spalten A-C sind nur ausgeblendet.",
        "SW 00 und SW 01 werden bewusst getrennt: Vorbereitung/Schulstart und erster regulaerer Unterricht.",
        "Ferien im Ferien-Tab pflegen. Ferien nicht zusaetzlich als eigene Terminplan-Zeilen eintragen, sonst entstehen Dubletten.",
        "Teilwochen: schulfreie Tage zu Beginn als eigene Termine eintragen, Kategorie Feiertage/Ferien, Ganztägig = Ja.",
    ]
    for note in notes:
        write(r, 2, ">> " + note)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = None
        r += 1

    section("Spalten")
    rows = [
        ("D Schulwoche", "Berechnet aus Ferien & Eckdaten."),
        ("E Wochentag", "Mo, Di, Mi, Do, Fr oder Ganze Woche."),
        ("F/G Uhrzeit", "Start und optional Ende, Format 08:30."),
        ("H Titel", "Pflichtfeld fuer Konverter."),
        ("I Kategorie", "Dropdown aus Kategorien."),
        ("J Ganztägig", "Ja oder Nein."),
        ("K Anmerkung", "Optional, bei SW-Headern auch Warnhinweise."),
    ]
    for left, right in rows:
        write(r, 2, left, bold=True, bg=C_ZEBRA)
        write(r, 3, right, bg=C_ZEBRA)
        r += 1

    return ws


def build_workbook() -> Workbook:
    wb = Workbook()
    del wb[wb.sheetnames[0]]
    make_terminplan(wb)
    make_ferien(wb)
    make_kategorien(wb)
    make_anleitung(wb)
    wb.active = 0
    return wb


def main():
    wb = build_workbook()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("OK - neue bearbeitbare Vorlage gespeichert:", OUT)
    print("Tabs:", ", ".join(wb.sheetnames))
    print("Schutz: keiner. Technische Spalten A-C im Terminplan sind ausgeblendet.")


if __name__ == "__main__":
    main()
