#!/usr/bin/env python3
"""
Schulwochen-Vorlagen-Generator
==============================

Erzeugt eine fertige Excel-Vorlage fuer das Schuljahr basierend auf der
Vorlage 'Terminplan_2026_27.xlsx'. Alle Formeln, Dropdowns und Formatierungen
bleiben erhalten.

Nutzung:
    python generate_school_template.py --config ferien.json
    python generate_school_template.py --interactive

Die Ausgabe ist eine voll funktionsfaehige Excel-Datei mit:
  - Automatischer Schulwochen-Berechnung (SW 00-41)
  - Ferien-ueberspringenden Montags-Datumsformeln
  - Dropdowns fuer Wochentag, Kategorie, Ganztaegig
  - Data Validation und Blattschutz
  - Vier Sheets: Terminplan, Ferien, Kategorien, Anleitung
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Fehler: openpyxl ist nicht installiert.")
    print("Installiere mit:  pip install openpyxl")
    sys.exit(1)


TEMPLATE_PATH = Path(__file__).parent.parent / "vorlagen" / "Terminplan_2026_27.xlsx"


def parse_date(s):
    """ISO-Datum (yyyy-mm-dd) -> date-Objekt."""
    if not s:
        return None
    if isinstance(s, date):
        return s
    y, m, d = map(int, s.split("-"))
    return date(y, m, d)


def load_config(path):
    """Lade Ferien-Konfiguration aus JSON-Datei."""
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg


def interactive_config():
    """Ferien-Daten interaktiv abfragen."""
    print("=" * 50)
    print("Schulwochen-Vorlagen-Generator")
    print("=" * 50)
    print()
    print("Bitte die Eckdaten fuer das neue Schuljahr eingeben.")
    print("Format fuer alle Daten: JJJJ-MM-TT  (z.B. 2026-08-24)")
    print()

    def ask(prompt, default=None):
        if default:
            val = input(f"{prompt} [{default}]: ").strip()
            return val or default
        return input(f"{prompt}: ").strip()

    cfg = {
        "schuljahr": ask("Schuljahr", "2026/27"),
        "sw00": ask("Erster Schultag (SW 00)", "2026-08-24"),
        "sw01": ask("Erster Unterrichtstag (SW 01)", "2026-08-31"),
        "last_day": ask("Letzter Schultag", "2027-07-16"),
        "holidays": [
            {
                "name": "Herbstferien",
                "start": ask("Herbstferien von", "2026-10-19"),
                "end": ask("Herbstferien bis", "2026-10-30"),
            },
            {
                "name": "Weihnachtsferien",
                "start": ask("Weihnachtsferien von", "2026-12-22"),
                "end": ask("Weihnachtsferien bis", "2027-01-03"),
            },
            {
                "name": "Osterferien",
                "start": ask("Osterferien von", "2027-03-22"),
                "end": ask("Osterferien bis", "2027-04-02"),
            },
            {
                "name": "Pfingstferien",
                "start": ask("Pfingstferien von (optional, Enter fuer leer)", ""),
                "end": ask("Pfingstferien bis (optional, Enter fuer leer)", ""),
            },
            {
                "name": "Sommerferien",
                "start": ask("Sommerferien von", "2027-07-19"),
                "end": ask("Sommerferien bis", "2027-08-31"),
            },
        ],
    }

    # Remove empty optional holidays
    cfg["holidays"] = [
        h for h in cfg["holidays"]
        if h["start"] or h["name"] != "Pfingstferien"
    ]

    save = ask("Konfiguration als JSON speichern? (j/n)", "n")
    if save.lower() == "j":
        filename = f"ferien_{cfg['schuljahr'].replace('/', '_')}.json"
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False, default=str)
        print(f"Gespeichert: {filename}")

    return cfg


def generate(cfg):
    """Erzeuge die Schulwochen-Vorlage."""
    if not TEMPLATE_PATH.exists():
        print(f"Fehler: Template nicht gefunden: {TEMPLATE_PATH}")
        sys.exit(1)

    print(f"Lade Template: {TEMPLATE_PATH}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)

    # ═══════════════════════════════════════════════════════════
    # 1. FERIEN-SHEET aktualisieren
    # ═══════════════════════════════════════════════════════════
    ws_ferien = wb["Ferien"]

    # Ferienzeilen: B3-B7 = Startdaten, C3-C7 = Enddaten
    holiday_rows = [3, 4, 5, 6, 7]  # Herbst, Weihnachten, Ostern, Pfingsten, Sommer
    for i, h in enumerate(cfg["holidays"]):
        row = holiday_rows[i]
        start = parse_date(h.get("start"))
        end = parse_date(h.get("end"))

        if start:
            ws_ferien.cell(row=row, column=2, value=start)
        else:
            ws_ferien.cell(row=row, column=2, value=None)

        if end:
            ws_ferien.cell(row=row, column=3, value=end)
        else:
            ws_ferien.cell(row=row, column=3, value=None)

        # Hinweis-Spalte aktualisieren
        hint = "Pflichtfeld" if start else "Optional leer lassen"
        ws_ferien.cell(row=row, column=4, value=hint)

    # Eckdaten
    ws_ferien.cell(row=10, column=2, value=parse_date(cfg["sw00"]))   # Erster Schultag
    ws_ferien.cell(row=11, column=2, value=parse_date(cfg["sw01"]))   # Erster Unterrichtstag
    ws_ferien.cell(row=12, column=2, value=parse_date(cfg["last_day"]))  # Letzter Schultag

    # ═══════════════════════════════════════════════════════════
    # 2. ANLEITUNG aktualisieren
    # ═══════════════════════════════════════════════════════════
    ws_anl = wb["Anleitung"]
    ws_anl.cell(row=2, column=2, value=f"Schuljahreskalender {cfg['schuljahr']} - Anleitung")

    # ═══════════════════════════════════════════════════════════
    # 3. KATEGORIEN aktualisieren (falls gewuenscht)
    # ═══════════════════════════════════════════════════════════
    if cfg.get("categories"):
        ws_kat = wb["Kategorien"]
        # Kategorien ab Zeile 3
        for i, cat in enumerate(cfg["categories"], start=3):
            ws_kat.cell(row=i, column=1, value=cat["label"])
            ws_kat.cell(row=i, column=2, value=cat.get("color", ""))
            ws_kat.cell(row=i, column=3, value=cat.get("keywords", ""))

    # ═══════════════════════════════════════════════════════════
    # 4. SPEICHERN
    # ═══════════════════════════════════════════════════════════
    safe_name = cfg["schuljahr"].replace("/", "_").replace("\\", "_")
    out_name = f"Terminplan_Schulwochen_Vorlage_{safe_name}.xlsx"
    out_path = Path(out_name)

    wb.save(out_path)
    print(f"\nErfolg! Vorlage gespeichert: {out_path.absolute()}")
    print(f"  - {ws_ferien.max_row} Zeilen im Ferien-Sheet")
    print(f"  - {wb['Terminplan'].max_row} Zeilen im Terminplan-Sheet")
    print(f"  - Alle Formeln, Dropdowns und Formatierungen erhalten")
    return out_path


def main():
    parser = argparse.ArgumentParser(
        description="Schulwochen-Vorlagen-Generator",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  python generate_school_template.py --interactive
  python generate_school_template.py --config ferien_2026_27.json
  python generate_school_template.py --config ferien.json --output meine_vorlage.xlsx

JSON-Format fuer --config:
  {
    "schuljahr": "2026/27",
    "sw00": "2026-08-24",
    "sw01": "2026-08-31",
    "last_day": "2027-07-16",
    "holidays": [
      {"name": "Herbstferien", "start": "2026-10-19", "end": "2026-10-30"},
      {"name": "Weihnachtsferien", "start": "2026-12-22", "end": "2027-01-03"},
      {"name": "Osterferien", "start": "2027-03-22", "end": "2027-04-02"},
      {"name": "Pfingstferien", "start": null, "end": null},
      {"name": "Sommerferien", "start": "2027-07-19", "end": "2027-08-31"}
    ]
  }
        """,
    )
    parser.add_argument("--config", "-c", help="Pfad zur JSON-Konfigurationsdatei")
    parser.add_argument("--interactive", "-i", action="store_true", help="Interaktiver Modus (Fragen werden gestellt)")
    parser.add_argument("--output", "-o", help="Ausgabedatei (optional)")
    args = parser.parse_args()

    if args.interactive:
        cfg = interactive_config()
    elif args.config:
        cfg = load_config(args.config)
    else:
        parser.print_help()
        print("\nFehler: Entweder --interactive oder --config angeben.")
        sys.exit(1)

    out_path = generate(cfg)

    if args.output:
        import shutil
        shutil.copy(out_path, args.output)
        print(f"Zusaetzlich kopiert nach: {args.output}")


if __name__ == "__main__":
    main()
