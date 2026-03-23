"""
patch_xlsx.py v2 – Dynamische Datumsformeln (kompatibel, ohne LET/FILTER/SEQUENCE)

Problem mit v1:
  - Zahlenformat "JJJJ-MM-TT" ist kein gültiger OOXML-Code → Datum als Zahl angezeigt
  - LET/FILTER/SEQUENCE benötigen _xlfn.-Präfix + Dynamic-Array-Markup → Datei-Reparatur

Lösung:
  1. Hilfstabelle Ferien!E14:F55 – 41 Schulmontage mit IF+WEEKDAY-Kette berechnet.
     Jede Zeile: vorheriger Montag + 7 Tage, Ferienwochen werden übersprungen.
     Funktioniert in Excel 365, LTSC 2021/2024 und älteren Versionen.
  2. Terminplan Spalte B: einfacher Bezug =Ferien!$F$15 … $F$55 (kein Kompatibilitätsproblem)
  3. Terminplan Spalte D: TEXT-Formel (Deutsch: TT/MM/JJJJ für German-Excel-Locale)
  4. Zahlenformate: OOXML-konform – YYYY-MM-DD für ISO, DD.MM.YYYY für Deutsch

Ausführen: python scripts/patch_xlsx.py
"""

import openpyxl
from openpyxl.styles import Font
from datetime import datetime

XLSX_PATH = 'website/downloads/GSH_Terminplan_Schulwochen_Vorlage.xlsx'

# Hilfstabelle im Ferien-Blatt: Spalten E (Label) und F (Datum)
# F15 = SW 00, F16 = SW 01, …, F56 = SW 41
HELPER_HEADER_ROW = 14
HELPER_START_ROW  = 15   # SW 00
HELPER_COL_LABEL  = 5    # Spalte E
HELPER_COL_DATE   = 6    # Spalte F


def parse_german_date(s):
    """TT.MM.JJJJ → datetime.date"""
    return datetime.strptime(s.strip(), '%d.%m.%Y').date()


def helper_formula_sw0():
    """F15: erster Schultag = $B$10"""
    return '=$B$10'


def helper_formula_sw_n(prev_row):
    """
    F{prev_row+1}: nächster Schulmontag nach $F${prev_row}.

    Logik:
      candidate = $F${prev_row} + 7
      Für jede Ferienperiode (B/C-Paar):
        Falls candidate im Ferienzeitraum liegt:
          springe auf ersten Montag NACH Ferienende
          = Ferienende + 8 - WEEKDAY(Ferienende, 2)
          Zusätzliche Tage = (erster Montag nach Ferien) - candidate
        Falls leer: 0 Zusatztage

    WEEKDAY(X, 2): Mo=1, Di=2, …, Sa=6, So=7
    Erster Montag nach X = X + 8 - WEEKDAY(X, 2)
    (gilt auch wenn X = Montag → gibt nächsten Montag +7 zurück, korrekt wenn Ferienletzter = Montag)
    """
    fp = f'$F${prev_row}'

    def skip(b, c):
        """Zusätzliche Tage falls Candidate in Ferienzeitraum [b, c] liegt."""
        return (
            f'+IF(ISBLANK({b}),0,'
            f'IF(({fp}+7>={b})*({fp}+7<={c}),'
            f'{c}+8-WEEKDAY({c},2)-({fp}+7),0))'
        )

    return (
        f'={fp}+7'
        + skip('$B$3', '$C$3')   # Herbstferien
        + skip('$B$4', '$C$4')   # Weihnachtsferien
        + skip('$B$5', '$C$5')   # Osterferien
        + skip('$B$6', '$C$6')   # Pfingstferien (darf leer sein)
    )


def terminplan_b_formula(helper_row):
    """Einfacher Bezug auf berechneten Schulmontag in der Ferien-Hilfstabelle."""
    return f'=Ferien!$F${helper_row}'


def terminplan_d_formula(row):
    """
    Wochenbezeichnung, z.B. "18.08. – 22.08.2025".
    TEXT() verwendet German-Locale-Codes (TT, MM, JJJJ) – korrekt für German Excel.
    En-Dash U+2013 als Trenner.
    """
    return f'=TEXT(B{row},"TT.MM.")&" \u2013 "&TEXT(B{row}+4,"TT.MM.JJJJ")'


# ---------------------------------------------------------------------------
# Datei laden
# ---------------------------------------------------------------------------
print(f'Lade: {XLSX_PATH}')
wb = openpyxl.load_workbook(XLSX_PATH)


# ===========================================================================
# Ferien-Blatt
# ===========================================================================
ws_f = wb['Ferien']

# --- 1. Datumszellen: Text → echter Datumswert (OOXML-konformes Format) ---
ferien_addr_pairs = [
    ('B3', 'C3'),   # Herbstferien
    ('B4', 'C4'),   # Weihnachtsferien
    ('B5', 'C5'),   # Osterferien
    ('B6', 'C6'),   # Pfingstferien (darf leer sein)
    ('B7', 'C7'),   # Sommerferien
]
for von_addr, bis_addr in ferien_addr_pairs:
    for addr in (von_addr, bis_addr):
        cell = ws_f[addr]
        if isinstance(cell.value, str) and cell.value.strip():
            cell.value = parse_german_date(cell.value)
            cell.number_format = 'DD.MM.YYYY'    # OOXML-Standard, nicht TT.MM.JJJJ

for addr in ('B10', 'B11', 'B12'):
    cell = ws_f[addr]
    if isinstance(cell.value, str) and cell.value.strip():
        cell.value = parse_german_date(cell.value)
        cell.number_format = 'DD.MM.YYYY'

# --- 2. A1: Schuljahr-Formel ---
ws_f['A1'].value = (
    '=IFERROR("Schuljahr "&TEXT(YEAR($B$10),"0")&"/"'
    '&TEXT(YEAR($B$10)+1-2000,"00")&" \u2013 Ferienliste","Ferienliste")'
)

# --- 3. Hilfstabelle aufräumen (idempotent bei mehrfachem Ausführen) ---
for r in range(HELPER_HEADER_ROW, HELPER_START_ROW + 43):
    ws_f.cell(r, HELPER_COL_LABEL).value = None
    ws_f.cell(r, HELPER_COL_DATE).value  = None

# --- 4. Hilfstabelle schreiben ---
# Header-Zeile (Zeile 14)
h = ws_f.cell(HELPER_HEADER_ROW, HELPER_COL_LABEL)
h.value = 'Berechnete Schulmontage \u2013 NICHT manuell \u00e4ndern'
h.font  = Font(bold=True, italic=True, color='888888', size=9)

# SW 00 (Zeile 15): = $B$10
cell0 = ws_f.cell(HELPER_START_ROW, HELPER_COL_DATE)
cell0.value = helper_formula_sw0()
cell0.number_format = 'DD.MM.YYYY'
ws_f.cell(HELPER_START_ROW, HELPER_COL_LABEL).value = 'SW 00'

# SW 01–41 (Zeilen 16–56): Kette
for sw_num in range(1, 42):
    row      = HELPER_START_ROW + sw_num
    prev_row = HELPER_START_ROW + sw_num - 1
    cell_d = ws_f.cell(row, HELPER_COL_DATE)
    cell_d.value = helper_formula_sw_n(prev_row)
    cell_d.number_format = 'DD.MM.YYYY'
    ws_f.cell(row, HELPER_COL_LABEL).value = f'SW {sw_num:02d}'

print(f'  Ferien-Blatt: Datumswerte, A1-Formel, Hilfstabelle F{HELPER_START_ROW}:F{HELPER_START_ROW+41} gesetzt')


# ===========================================================================
# Terminplan-Blatt
# ===========================================================================
ws_t = wb['Terminplan']

# SW-Header-Zeilen ermitteln
sw_header_rows = []
for row in range(2, ws_t.max_row + 1):
    val_a = ws_t.cell(row, 1).value
    if val_a and isinstance(val_a, str) and val_a.startswith('SW ') and val_a != 'SW-Key':
        try:
            sw_num = int(val_a.split()[1])
            sw_header_rows.append((sw_num, row))
        except (IndexError, ValueError):
            pass

print(f'  Terminplan: {len(sw_header_rows)} SW-Header-Zeilen gefunden')

for sw_num, row in sw_header_rows:
    helper_row = HELPER_START_ROW + sw_num   # SW 00 → F15, SW 01 → F16, …

    # Spalte B: Bezug auf Ferien-Hilfstabelle (gibt Datums-Seriennummer zurück)
    cell_b = ws_t.cell(row, 2)
    cell_b.value = terminplan_b_formula(helper_row)
    cell_b.number_format = 'YYYY-MM-DD'     # OOXML-Standard für ISO-Datum

    # Spalte D: Wochenbezeichnung als Text-Formel
    cell_d = ws_t.cell(row, 4)
    cell_d.value = terminplan_d_formula(row)
    # number_format bleibt General (TEXT() gibt bereits einen String zurück)

print(f'  Terminplan: Formeln B+D fuer SW {sw_header_rows[0][0]:02d} (Z.{sw_header_rows[0][1]}) '
      f'bis SW {sw_header_rows[-1][0]:02d} (Z.{sw_header_rows[-1][1]}) gesetzt')


# ---------------------------------------------------------------------------
# Speichern
# ---------------------------------------------------------------------------
wb.save(XLSX_PATH)
print(f'Gespeichert: {XLSX_PATH}')
print('Fertig. Bitte recalc.py zur Verifikation ausfuehren.')
