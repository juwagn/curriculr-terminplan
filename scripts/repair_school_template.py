# -*- coding: utf-8 -*-
"""
Repariert die Terminplan-Arbeitsmappen fuer den wiederverwendbaren Schuljahresprozess.

Ziele:
- Spalte B im Terminplan wird aus Ferien!F15:F56 abgeleitet, nicht als Text fixiert.
- Ferien werden nur im Ferien-Blatt gepflegt; exakte Ferien-Platzhalter im Terminplan
  werden geleert, damit der Konverter sie nicht doppelt importiert.
- Dropdowns gelten bis zur letzten Datenzeile.
- Blattschutz schuetzt technische Bereiche; echte Eingabefelder bleiben editierbar.

Ausfuehren: python scripts/repair_school_template.py
"""

from pathlib import Path
import re
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation


BASE = Path(__file__).resolve().parent.parent
FILES = [
    BASE / "vorlagen" / "Terminplan_2026_27.xlsx",
    BASE / "website" / "downloads" / "Terminplan_Schulwochen_Vorlage.xlsx",
]

SW_COUNT = 42
HELPER_START_ROW = 15
INPUT_PASSWORD = "school"


def is_sw_header(value):
    text = str(value or "").strip()
    match = re.match(r"^SW\s+(\d{1,2})$", text)
    if not match:
        return None
    sw_num = int(match.group(1))
    return sw_num if 0 <= sw_num < SW_COUNT else None


def remove_existing_validations(ws):
    ws.data_validations.dataValidation = []


def add_validations(ws):
    max_row = ws.max_row
    dv_weekday = DataValidation(
        type="list",
        formula1='"Mo,Di,Mi,Do,Fr,Ganze Woche"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Ungueltige Eingabe",
        error="Bitte einen Wochentag aus dem Dropdown waehlen.",
    )
    dv_weekday.add(f"E2:E{max_row}")
    ws.add_data_validation(dv_weekday)

    dv_category = DataValidation(
        type="list",
        formula1="KategorienListe",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Ungueltige Kategorie",
        error="Bitte eine Kategorie aus dem Dropdown waehlen.",
    )
    dv_category.add(f"I2:I{max_row}")
    ws.add_data_validation(dv_category)

    dv_allday = DataValidation(
        type="list",
        formula1='"Ja,Nein"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Ungueltige Eingabe",
        error='Bitte "Ja" oder "Nein" auswaehlen.',
    )
    dv_allday.add(f"J2:J{max_row}")
    ws.add_data_validation(dv_allday)


def holiday_names(ws_ferien):
    names = set()
    for row in range(3, 8):
        name = str(ws_ferien.cell(row, 1).value or "").strip().lower()
        if name:
            names.add(name)
    return names


def clear_duplicate_holidays(ws, names):
    cleared = []
    for row in range(2, ws.max_row + 1):
        title = str(ws.cell(row, 8).value or "").strip().lower()
        if title in names:
            for col in range(5, 12):
                ws.cell(row, col).value = None
            cleared.append(row)
    return cleared


def apply_school_week_formulas(ws):
    current_sw = None
    header_count = 0
    data_count = 0

    for row in range(2, ws.max_row + 1):
        sw_num = is_sw_header(ws.cell(row, 1).value)
        if sw_num is not None:
            current_sw = sw_num
            ws.cell(row, 2).value = f"=Ferien!$F${HELPER_START_ROW + sw_num}"
            ws.cell(row, 2).number_format = "YYYY-MM-DD"
            ws.cell(row, 4).value = f'=TEXT(B{row},"TT.MM.")&" - "&TEXT(B{row}+4,"TT.MM.JJJJ")'
            header_count += 1
            continue

        if current_sw is not None:
            ws.cell(row, 2).value = f"=Ferien!$F${HELPER_START_ROW + current_sw}"
            ws.cell(row, 2).number_format = "YYYY-MM-DD"
            data_count += 1

    return header_count, data_count


def protect_workbook(wb):
    ws_t = wb["Terminplan"]
    ws_f = wb["Ferien"]
    ws_k = wb["Kategorien"]

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=True)

    for row in range(2, ws_t.max_row + 1):
        if is_sw_header(ws_t.cell(row, 1).value) is None:
            for col in range(5, 12):
                ws_t.cell(row, col).protection = Protection(locked=False)

    for addr in ["B3:C7", "B10:B12"]:
        for row in ws_f[addr]:
            for cell in row:
                cell.protection = Protection(locked=False)

    for row in range(3, max(ws_k.max_row + 1, 20)):
        for col in range(1, 4):
            ws_k.cell(row, col).protection = Protection(locked=False)

    for ws in wb.worksheets:
        ws.protection.sheet = True
        ws.protection.password = INPUT_PASSWORD
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = True


def repair_file(path):
    if not path.exists():
        return {"file": str(path), "status": "missing"}

    backup = path.with_suffix(path.suffix + ".repair-bak")
    shutil.copy(path, backup)

    wb = load_workbook(path)
    if "Terminplan" not in wb.sheetnames or "Ferien" not in wb.sheetnames:
        return {"file": str(path), "status": "error", "error": "Terminplan/Ferien fehlt"}

    ws_t = wb["Terminplan"]
    ws_f = wb["Ferien"]

    headers, data_rows = apply_school_week_formulas(ws_t)
    cleared = clear_duplicate_holidays(ws_t, holiday_names(ws_f))

    remove_existing_validations(ws_t)
    add_validations(ws_t)
    protect_workbook(wb)

    wb.save(path)
    return {
        "file": str(path),
        "status": "ok",
        "backup": str(backup),
        "sw_headers": headers,
        "data_rows": data_rows,
        "cleared_duplicate_holiday_rows": cleared,
    }


def main():
    for file_path in FILES:
        result = repair_file(file_path)
        print(result)


if __name__ == "__main__":
    main()
