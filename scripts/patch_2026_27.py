"""
patch_2026_27.py – Einmaliger Patch fuer Terminplan_2026_27.xlsx

Behebt:
- Spalte B in SW-Header-Zeilen: Formelstrings -> ISO-Datumswerte
- Spalte B in Datenzeilen: None -> ISO-Datumswert des Schulwoche-Montags

Berechnet alle 42 Schulmontage (SW 00–41) direkt aus dem Ferien-Sheet
(B10 = erster Schultag, B3:C7 = Ferienperioden). Keine Abhaengigkeit von
gecachten Excel-Formelwerten.

Ausfuehren: python scripts/patch_2026_27.py
"""

import re
import shutil
from datetime import date, timedelta
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
SRC  = BASE / 'vorlagen' / 'Terminplan_2026_27.xlsx'
OUT  = BASE / 'vorlagen' / 'Terminplan_2026_27.xlsx'
BAK  = BASE / 'vorlagen' / 'Terminplan_2026_27.xlsx.bak'

SW_COUNT = 42  # SW 00 bis SW 41


def lese_ferien(ws_ferien):
    """Liest Ferienperioden und Schuljahresstart aus dem Ferien-Sheet."""
    def to_date(val):
        if val is None:
            return None
        if hasattr(val, 'date'):
            return val.date()
        return None

    schuljahr_start = to_date(ws_ferien.cell(10, 2).value)
    assert schuljahr_start is not None, 'B10 (erster Schultag) fehlt im Ferien-Sheet'

    perioden = []
    for row in range(3, 8):  # B3:C7
        von = to_date(ws_ferien.cell(row, 2).value)
        bis = to_date(ws_ferien.cell(row, 3).value)
        if von and bis:
            perioden.append((von, bis))

    return schuljahr_start, perioden


def berechne_schulmontage(schuljahr_start, ferien_perioden):
    """Berechnet SW_COUNT Schulmontage ab Schuljahresstart, Ferienwochen ueberspringen."""
    def ist_ferien(monday):
        for von, bis in ferien_perioden:
            if von <= monday <= bis:
                return True
        return False

    mondays = []
    current = schuljahr_start
    while len(mondays) < SW_COUNT:
        if not ist_ferien(current):
            mondays.append(current)
        current += timedelta(weeks=1)

    return mondays


def patch_terminplan(ws_terminplan, mondays):
    """Schreibt ISO-Datumswerte in Spalte B aller SW-Header- und Datenzeilen."""
    current_monday = None
    patched_headers = 0
    patched_data = 0

    for row in range(1, ws_terminplan.max_row + 1):
        a_val = ws_terminplan.cell(row, 1).value
        a_str = str(a_val).strip() if a_val is not None else ''

        # SW-Header-Zeile: A = "SW 00" … "SW 41"
        match = re.match(r'^SW\s+(\d{1,2})$', a_str)
        if match:
            sw_num = int(match.group(1))
            if 0 <= sw_num < len(mondays):
                current_monday = mondays[sw_num]
                ws_terminplan.cell(row, 2).value = current_monday.isoformat()
                patched_headers += 1
            continue

        # Datenzeile: A leer
        if not a_str and current_monday is not None:
            ws_terminplan.cell(row, 2).value = current_monday.isoformat()
            patched_data += 1

    return patched_headers, patched_data


def main():
    assert SRC.exists(), f'Quelldatei nicht gefunden: {SRC}'

    # Backup erstellen
    shutil.copy(SRC, BAK)
    print(f'Backup: {BAK}')

    wb = openpyxl.load_workbook(str(SRC))

    assert 'Ferien' in wb.sheetnames, 'Ferien-Sheet fehlt'
    assert 'Terminplan' in wb.sheetnames, 'Terminplan-Sheet fehlt'

    schuljahr_start, ferien = lese_ferien(wb['Ferien'])
    print(f'Schuljahresstart: {schuljahr_start}')
    print(f'Ferienperioden:   {len(ferien)} gefunden')

    mondays = berechne_schulmontage(schuljahr_start, ferien)
    print(f'Schulmontage:     SW 00={mondays[0]}  SW 41={mondays[41]}')

    h, d = patch_terminplan(wb['Terminplan'], mondays)
    print(f'Spalte B gesetzt: {h} SW-Header-Zeilen, {d} Datenzeilen')

    wb.save(str(OUT))
    print(f'\nOK – gespeichert: {OUT}')


if __name__ == '__main__':
    main()
